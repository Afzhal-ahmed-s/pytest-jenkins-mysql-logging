import requests
import json
import jsonpath
import openpyxl

def test_add_multiple_students():

    # API
    API_URL = "http://thetestingworldapi.com/api/studentsDetails"
    file = open("/home/afzhal-ahmed-s/PycharmProjects/AddNewStudent.json")
    json_request = json.loads(file.read())

    # Excel code
    wb = openpyxl.load_workbook("/home/afzhal-ahmed-s/Downloads/Noduco_Udemy_C-PPA.xlsx")
    sh = wb['Sheet1']
    rows = sh.max_row

    for i in range(2, rows + 1):
        cell_first_name = sh.cell(row = i, column=1)
        cell_mid_name = sh.cell(row = i, column=2)
        cell_last_name = sh.cell(row = i, column=3)
        cell_dob = sh.cell(row = i, column=4)

        json_request['first_name'] = cell_first_name.value
        json_request['middle_name'] = cell_mid_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_dob.value

        response = requests.post(API_URL, json_request)

        print(response.status_code)
        print(json_request)

        assert response.status_code == 200

