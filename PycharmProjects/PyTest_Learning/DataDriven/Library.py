import json
import jsonpath
import openpyxl
import requests

class Common:

    def __init__(self, FileNamePath, SheetName):
        global wk
        global sh
        wb = openpyxl.load_workbook(FileNamePath)
        sh = wb[SheetName]


    def fetch_row_count(self):
        # Excel code
        rows = sh.max_row
        return rows

    def fetch_column_count(self):
        # Excel code
        cols = sh.max_column
        return cols

    def fetch_key_names(self):
        c = sh.max_column
        li = []
        for i in range(1, c + 1):
            cell = sh.cell(row=1, column=i)
            li.insert(i-1, cell.value)

        return li

    def update_request_with_data(self, rowNumber, jsonRequest, keyList):
        c = sh.max_column
        for i in range(1, c+1):
            cell = sh.cell(rowNumber, column=i)
            jsonRequest[keyList[i-1]] = cell.value

        return jsonRequest
