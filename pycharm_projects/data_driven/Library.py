import json
import jsonpath
import openpyxl
import requests

class Common:

    def __init__(self, file_name_path, sheet_name):
        global wb
        global sh
        global file_path
        file_path = file_name_path
        wb = openpyxl.load_workbook(file_name_path)
        sh = wb[sheet_name]


    def fetch_row_count(self):
        # Excel code
        rows = sh.max_row
        return rows

    def fetch_column_count(self):
        # Excel code
        cols = sh.max_column
        return cols

    def fetch_key_names(self):
        c = sh.max_column
        li = []
        for i in range(1, c + 1):
            cell = sh.cell(row=1, column=i)
            li.insert(i-1, cell.value)

        return li

    def update_request_with_data(self, row_number, json_request, key_list):
        c = sh.max_column
        for i in range(1, c+1):
            cell = sh.cell(row_number, column=i)
            json_request[key_list[i - 1]] = cell.value

        return json_request

    def set_column_data(self, row_number, number_of_columns, list_of_data):
        for i in range(1, number_of_columns + 1):
            cell = sh.cell(row= row_number, column= i)
            cell.value = list_of_data[i -1]
            print("iterating row number: ", row_number, "column number: ", i)

        wb.save(file_path)